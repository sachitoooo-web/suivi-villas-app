import openpyxl
import re
import datetime

def scanner_planning_excel(fichier_upload):
    wb = openpyxl.load_workbook(fichier_upload, data_only=True)
    projets_dict = {}
    
    regex_prs = re.compile(r'(PRS\d{5,7})', re.IGNORECASE)
    onglets_a_scanner = [nom for nom in wb.sheetnames if '2025' in nom or '2026' in nom]
    
    for nom_onglet in onglets_a_scanner:
        ws = wb[nom_onglet]
        
        # 1. Cartographier les dates des colonnes (en-têtes)
        ligne_des_dates = None
        dates_colonnes = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False)):
            for col_idx, cell in enumerate(row):
                if isinstance(cell.value, datetime.datetime):
                    ligne_des_dates = i + 1
                    break
            if ligne_des_dates:
                for col_idx, cell in enumerate(row):
                    if isinstance(cell.value, datetime.datetime):
                        dates_colonnes[col_idx] = cell.value.date()
                break
                
        if not ligne_des_dates:
            continue
            
        # 2. Scanner toutes les cellules de l'onglet
        for row_idx, row in enumerate(ws.iter_rows(min_row=ligne_des_dates + 1, values_only=False), start=ligne_des_dates + 1):
            for col_idx, cell in enumerate(row):
                if cell.value and isinstance(cell.value, str):
                    match_prs = regex_prs.search(cell.value)
                    
                    if match_prs:
                        prs_code = match_prs.group(1).upper()
                        
                        # Initialiser le projet s'il est nouveau
                        if prs_code not in projets_dict:
                            projets_dict[prs_code] = {
                                "nom": str(cell.value).replace('\n', ' ').strip(),
                                "is_sacha": False,
                                "dates_jaunes": [],
                                "dates_toutes": []
                            }
                        
                        # Récupérer la date de la colonne actuelle
                        cell_date = dates_colonnes.get(col_idx)
                        if cell_date:
                            projets_dict[prs_code]["dates_toutes"].append(cell_date)
                            
                        # Vérifier la couleur de la case
                        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                            couleur_hex = str(cell.fill.start_color.rgb).upper()
                            
                            # Détection du VERT (Tes projets) - Exclut le bleu/cyan des autres CDP
                            verts_acceptes = ['00B050', '28A745', '39B54A', '32CD32', '00FF00', '92D050']
                            if any(vert in couleur_hex for vert in verts_acceptes):
                                projets_dict[prs_code]["is_sacha"] = True
                            elif len(couleur_hex) == 8 and couleur_hex[2:4] < '88' and couleur_hex[4:6] > 'A0' and couleur_hex[6:8] < '88':
                                projets_dict[prs_code]["is_sacha"] = True
                                
                            # Détection du JAUNE (Date de début)
                            if 'FFFF00' in couleur_hex or 'FFC000' in couleur_hex:
                                if cell_date:
                                    projets_dict[prs_code]["dates_jaunes"].append(cell_date)

    # 3. Consolidation et nettoyage final
    projets_finaux = []
    
    # On met une limite plus large pour commencer (1er Janvier 2025) pour être sûr de bien tout récupérer
    date_limite = datetime.date(2025, 1, 1) 
    
    for prs, data in projets_dict.items():
        if data["is_sacha"]:
            # On prend la date jaune en priorité, sinon la première date trouvée
            if data["dates_jaunes"]:
                date_debut = min(data["dates_jaunes"])
            elif data["dates_toutes"]:
                date_debut = min(data["dates_toutes"])
            else:
                continue # Si le projet n'a aucune date, on l'ignore
                
            if date_debut >= date_limite:
                projets_finaux.append({
                    "prs": prs,
                    "nom": data["nom"],
                    "date_debut": date_debut.isoformat(),
                    "cdp": "Sacha"
                })
            
    return projets_finaux